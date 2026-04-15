#!/usr/bin/env python3
"""
compare_points_interactive.py
==============================
Interactive script for comparing two worksheets inside one Excel workbook.
Prompts for a workbook, two sheet names, and comparison settings, then
writes the result sheets back into the same Excel file.

Supported input formats:
  - Excel (.xlsx, .xlsm)

Result categories:
  MATCH          - name and coordinates match
  NAME_CHANGED   - same coordinates, different name
  COORD_CHANGED  - same name, different coordinates -> show dX/dY/dZ
  DELETED        - exists only in the original sheet
  ADDED          - exists only in the new sheet
"""

import sys
import shutil
from copy import copy
from pathlib import Path
from datetime import datetime
from tempfile import NamedTemporaryFile
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ══════════════════════════════════════════════════════════════════════════════
#  DEFAULT SETTINGS
# ══════════════════════════════════════════════════════════════════════════════
DEFAULTS = {
    "coord_tol":       0.05,   # XYZ tolerance, mm
    "ijk_tol":         0.001,  # IJK tolerance
    "compare_ijk":     False,  # include I/J/K in coordinate comparison?
    "close_points_tol": 10.0,  # 3D distance threshold for close-points report, mm
}

PALETTE = {
    "MATCH":         "C6EFCE",
    "NAME_CHANGED":  "FFEB9C",
    "COORD_CHANGED": "FFD7D7",
    "MOVED":         "E2EFDA",
    "DELETED":       "F4CCCC",
    "ADDED":         "D9EAD3",
    "DIFF_CELL":     "CC3300",
    "HEADER_BG":     "4472C4",
    "HEADER_FG":     "FFFFFF",
    "TITLE":         "1F3864",
}

STATUS_LABEL = {
    "MATCH":         "✔  Match",
    "NAME_CHANGED":  "✎  Name Changed",
    "COORD_CHANGED": "⚠  Coordinates Changed",
    "MOVED":         "↔  Moved",
    "DELETED":       "✖  Deleted (only in original)",
    "ADDED":         "＋ Added (only in new)",
}

LOGICAL_FIELDS = ["Name", "X", "Y", "Z", "I", "J", "K"]
FONT_NAME = "Calibri"
REPORT_PREFIX = "CMP_"
REPORT_SUFFIXES = [
    "Overview",
    "All Results",
    "Match",
    "Name Changed",
    "Coord Changed",
    "Moved",
    "Deleted",
    "Added",
    "Close Points",
]

GROUP_HEADER_LABELS = {"Original data", "New Data", "Difference", "STATUS", "Comparison \nstatus"}
STATUS_GROUP_LABEL = "Comparison \nstatus"
LIGHT_THEME_COLOR = Color(theme=9, tint=0.7999816888943144)
GROUP_FILL_ORIGINAL = PatternFill(patternType="solid", fgColor=copy(LIGHT_THEME_COLOR))
GROUP_FILL_STATUS = PatternFill(patternType="solid", fgColor="FF5B9BD5")
GROUP_FILL_NEW = PatternFill(patternType="solid", fgColor="FFFFFF00")
GROUP_FILL_DIFF = PatternFill(patternType="solid", fgColor="FF5B9BD5")
STATUS_FILLS = {
    "MATCH": PatternFill(patternType="solid", fgColor=copy(LIGHT_THEME_COLOR)),
    "NAME_CHANGED": PatternFill(patternType="solid", fgColor="FFFFFF00"),
    "COORD_CHANGED": PatternFill(patternType="solid", fgColor="FFFFFF00"),
    "DELETED": PatternFill(patternType="solid", fgColor="FFFF0000"),
    "NEW": PatternFill(patternType="solid", fgColor="FFFFC000"),
}

ORIGINAL_SHEET_APPEND_MAP = [
    ("Status", "STATUS"),
    ("NEW_Name", "Name"),
    ("NEW_X", "X"),
    ("NEW_Y", "Y"),
    ("NEW_Z", "Z"),
    ("NEW_I", "I"),
    ("NEW_J", "J"),
    ("NEW_K", "K"),
    ("NAME_diff", "NAME_diff"),
    ("X_diff", "X_diff"),
    ("Y_diff", "Y_diff"),
    ("Z_diff", "Z_diff"),
    ("I_diff", "I_diff"),
    ("J_diff", "J_diff"),
    ("K_diff", "K_diff"),
    ("DIFF_Fields", "DIFF_Fields"),
]
LEGACY_ORIGINAL_SHEET_APPEND_HEADERS = [
    "STATUS",
    "NEW_Name",
    "NEW_X",
    "NEW_Y",
    "NEW_Z",
    "NEW_I",
    "NEW_J",
    "NEW_K",
    "NAME_diff",
    "X_diff",
    "Y_diff",
    "Z_diff",
    "I_diff",
    "J_diff",
    "K_diff",
    "DIFF_Fields",
]

ORIGINAL_SHEET_NUMERIC_COLUMNS = {
    "NEW_X", "NEW_Y", "NEW_Z", "NEW_I", "NEW_J", "NEW_K",
    "X_diff", "Y_diff", "Z_diff", "I_diff", "J_diff", "K_diff",
}
COORD_FIELDS = ["X", "Y", "Z", "I", "J", "K"]
COORD_NUMBER_FORMAT = "0.000"

# ══════════════════════════════════════════════════════════════════════════════
#  UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def hr(char="─", n=60):
    print(char * n)

def ask(prompt, default=None, validator=None):
    """Prompt the user and return a string."""
    hint = f" [{default}]" if default is not None else ""
    while True:
        raw = input(f"{prompt}{hint}: ").strip()
        if raw == "" and default is not None:
            return str(default)
        if raw == "" and default is None:
            print("  ⚠  A value is required.")
            continue
        if validator:
            ok, msg = validator(raw)
            if not ok:
                print(f"  ⚠  {msg}")
                continue
        return raw

def ask_yn(prompt, default=True):
    hint = "Y/n" if default else "y/N"
    raw = input(f"{prompt} [{hint}]: ").strip().lower()
    if raw == "":
        return default
    return raw in ("y", "yes")

def ask_float(prompt, default):
    while True:
        raw = input(f"{prompt} [{default}]: ").strip()
        if raw == "":
            return default
        try:
            return float(raw)
        except ValueError:
            print("  ⚠  Enter a numeric value such as 0.05")

def validate_workbook(path_str):
    p = Path(path_str)
    if not p.exists():
        return False, f"File not found: {p}"
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        return False, "Only .xlsx and .xlsm Excel files are supported"
    return True, ""

def list_sheets(path):
    """Return the worksheet names from an Excel file."""
    try:
        xl = pd.ExcelFile(path)
        return xl.sheet_names
    except Exception:
        return []

def read_sheet(path, sheet):
    """Read an Excel sheet into a DataFrame."""
    header_row = detect_header_row(path, sheet)
    return pd.read_excel(path, sheet_name=sheet, header=header_row - 1)

def detect_header_row(path, sheet):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        row1 = ["" if cell.value is None else str(cell.value).strip() for cell in ws[1]]
        row2 = ["" if cell.value is None else str(cell.value).strip() for cell in ws[2]]
    finally:
        wb.close()

    if any(value in GROUP_HEADER_LABELS for value in row1) and all(field in row2 for field in LOGICAL_FIELDS):
        return 2
    return 1


# ══════════════════════════════════════════════════════════════════════════════
#  COMPARISON LOGIC
# ══════════════════════════════════════════════════════════════════════════════

def prepare_sheet(df, sheet_name):
    """Validate required columns and normalize the sheet."""
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    missing = [field for field in LOGICAL_FIELDS if field not in df.columns]
    if missing:
        missing_str = ", ".join(missing)
        raise ValueError(
            f"Required columns are missing on sheet '{sheet_name}': {missing_str}. "
            "Fix the worksheet headers in the Excel file and run the tool again."
        )

    out = {}
    for field in LOGICAL_FIELDS:
        if field == "Name":
            out[field] = df[field].where(df[field].notna(), "").astype(str).str.strip()
        else:
            out[field] = pd.to_numeric(df[field], errors="coerce")
    prepared = pd.DataFrame(out)
    keep_mask = (prepared["Name"] != "") | prepared[["X", "Y", "Z", "I", "J", "K"]].notna().any(axis=1)
    return prepared.loc[keep_mask].reset_index(drop=True)

def coord_key(row, tol, ijk_tol, use_ijk):
    def rnd(v, t):
        try:
            return int(round(float(v) / t))
        except (TypeError, ValueError):
            return None
    k = (rnd(row["X"], tol), rnd(row["Y"], tol), rnd(row["Z"], tol))
    if use_ijk:
        k += (rnd(row["I"], ijk_tol), rnd(row["J"], ijk_tol), rnd(row["K"], ijk_tol))
    return k

def safe_delta(v1, v2):
    try:
        return round(float(v2) - float(v1), 4)
    except (TypeError, ValueError):
        return np.nan

def fmt_val(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "–"
    if isinstance(v, float):
        return f"{v:.4f}"
    return str(v)

def compare(df_orig, df_new, tol, ijk_tol, use_ijk):
    name_to_new  = {r["Name"]: i for i, r in df_new.iterrows()}
    coord_to_new = {}
    for i, r in df_new.iterrows():
        k = coord_key(r, tol, ijk_tol, use_ijk)
        coord_to_new.setdefault(k, []).append(i)

    matched_new = set()
    results = []

    for _, r1 in df_orig.iterrows():
        n1   = r1["Name"]
        ck1  = coord_key(r1, tol, ijk_tol, use_ijk)
        by_name  = n1 in name_to_new
        by_coord = ck1 in coord_to_new

        row = {f"ORIG_{f}": r1[f] for f in LOGICAL_FIELDS}
        row.update({f"NEW_{f}": np.nan for f in LOGICAL_FIELDS})
        row.update({
            "NAME_diff":  "",
            "X_diff": np.nan, "Y_diff": np.nan, "Z_diff": np.nan,
            "I_diff": np.nan, "J_diff": np.nan, "K_diff": np.nan,
            "DIFF_Fields": "",
        })

        if by_name:
            i2 = name_to_new[n1]
            r2 = df_new.loc[i2]
            matched_new.add(i2)
            row.update({f"NEW_{f}": r2[f] for f in LOGICAL_FIELDS})
            ck2 = coord_key(r2, tol, ijk_tol, use_ijk)
            if ck1 == ck2:
                row["Status"] = "MATCH"
            else:
                row["Status"] = "COORD_CHANGED"
                diffs = []
                for f in ["X","Y","Z","I","J","K"]:
                    d = safe_delta(r1[f], r2[f])
                    row[f"{f}_diff"] = d
                    if not (isinstance(d, float) and np.isnan(d)) and abs(d) > 1e-9:
                        diffs.append(f)
                row["DIFF_Fields"] = ", ".join(diffs)

        elif by_coord:
            found = False
            for i2 in coord_to_new[ck1]:
                if i2 not in matched_new:
                    r2 = df_new.loc[i2]
                    matched_new.add(i2)
                    row.update({f"NEW_{f}": r2[f] for f in LOGICAL_FIELDS})
                    row["Status"]    = "NAME_CHANGED"
                    row["NAME_diff"] = f"{n1}  →  {r2['Name']}"
                    row["DIFF_Fields"] = "Name"
                    found = True
                    break
            if not found:
                row["Status"] = "DELETED"
        else:
            row["Status"] = "DELETED"

        results.append(row)

    for i2, r2 in df_new.iterrows():
        if i2 not in matched_new:
            row = {"Status": "ADDED"}
            row.update({f"ORIG_{f}": np.nan for f in LOGICAL_FIELDS})
            row.update({f"NEW_{f}": r2[f] for f in LOGICAL_FIELDS})
            row.update({
                "NAME_diff": "", "DIFF_Fields": "",
                "X_diff": np.nan, "Y_diff": np.nan, "Z_diff": np.nan,
                "I_diff": np.nan, "J_diff": np.nan, "K_diff": np.nan,
            })
            results.append(row)

    return pd.DataFrame(results)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL REPORT
# ══════════════════════════════════════════════════════════════════════════════

COL_ORDER = [
    "Status",
    "ORIG_Name","ORIG_X","ORIG_Y","ORIG_Z","ORIG_I","ORIG_J","ORIG_K",
    "NEW_Name", "NEW_X", "NEW_Y", "NEW_Z", "NEW_I", "NEW_J", "NEW_K",
    "NAME_diff","X_diff","Y_diff","Z_diff","I_diff","J_diff","K_diff",
    "DIFF_Fields",
]

DIFF_COLS = {"NAME_diff","X_diff","Y_diff","Z_diff","I_diff","J_diff","K_diff","DIFF_Fields"}

def thin(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def ap(cell, style):
    for k, v in style.items():
        setattr(cell, k, v)

def hdr_style():
    return dict(
        font=Font(name=FONT_NAME, bold=True, color=PALETTE["HEADER_FG"], size=11),
        fill=PatternFill("solid", fgColor=PALETTE["HEADER_BG"]),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=thin(),
    )

def data_style(status, is_diff=False):
    bg = PALETTE.get(status, "FFFFFF")
    if is_diff:
        font = Font(name=FONT_NAME, size=10, bold=True, color=PALETTE["DIFF_CELL"])
    else:
        font = Font(name=FONT_NAME, size=10)
    return dict(
        font=font,
        fill=PatternFill("solid", fgColor=bg),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin(),
    )

def clone_font(base_font, **updates):
    new_font = copy(base_font)
    for key, value in updates.items():
        setattr(new_font, key, value)
    return new_font

def clone_alignment(base_alignment, **updates):
    new_alignment = copy(base_alignment)
    for key, value in updates.items():
        setattr(new_alignment, key, value)
    return new_alignment

def copy_style_or_default(value, default_factory):
    return copy(value) if value is not None else default_factory()

def find_first_non_empty_in_row(ws, row_idx, start_col, end_col):
    for col_idx in range(start_col, end_col + 1):
        if ws.cell(row_idx, col_idx).value not in (None, ""):
            return ws.cell(row_idx, col_idx)
    return None

def find_first_non_empty_data_cell(ws, data_start_row, source_end_col):
    for row_idx in range(data_start_row, ws.max_row + 1):
        for col_idx in range(1, source_end_col + 1):
            if ws.cell(row_idx, col_idx).value not in (None, ""):
                return ws.cell(row_idx, col_idx)
    return None

def infer_original_sheet_style_kit(ws, header_row, data_start_row, source_end_col):
    header_sample = find_first_non_empty_in_row(ws, header_row, 1, source_end_col)
    data_sample = find_first_non_empty_data_cell(ws, data_start_row, source_end_col)

    if header_sample is None:
        header_sample = ws.cell(header_row, 1)
    if data_sample is None:
        data_sample = ws.cell(data_start_row, 1)

    return {
        "header_font": copy_style_or_default(header_sample.font, lambda: Font(name=FONT_NAME, size=11, bold=True)),
        "header_fill": copy_style_or_default(header_sample.fill, lambda: PatternFill(patternType="solid", fgColor="FFFFFF")),
        "header_alignment": copy_style_or_default(header_sample.alignment, lambda: Alignment(horizontal="center", vertical="center")),
        "header_border": copy_style_or_default(header_sample.border, thin),
        "data_font": copy_style_or_default(data_sample.font, lambda: Font(name=FONT_NAME, size=11)),
        "data_alignment": copy_style_or_default(data_sample.alignment, lambda: Alignment(horizontal="left", vertical="center")),
        "data_border": copy_style_or_default(data_sample.border, thin),
    }

def original_sheet_header_style(style_kit, fill=None):
    return dict(
        font=clone_font(style_kit["header_font"], color="FF222222", bold=True, italic=False),
        fill=copy(fill) if fill is not None else PatternFill(patternType="solid", fgColor="FFEDEDED"),
        alignment=clone_alignment(style_kit["header_alignment"], horizontal="center", vertical="center", wrap_text=True),
        border=copy(style_kit["header_border"]),
    )

def original_sheet_group_style(style_kit, label):
    if label == "Original data":
        fill = GROUP_FILL_ORIGINAL
        size = 16
    elif label == STATUS_GROUP_LABEL:
        fill = GROUP_FILL_STATUS
        size = 11
    elif label == "New Data":
        fill = GROUP_FILL_NEW
        size = 16
    else:
        fill = GROUP_FILL_DIFF
        size = 16

    return dict(
        font=clone_font(style_kit["header_font"], size=size, bold=True, italic=False),
        fill=copy(fill),
        alignment=clone_alignment(style_kit["header_alignment"], horizontal="center", vertical="center", wrap_text=True),
        border=copy(style_kit["header_border"]),
    )

def original_sheet_data_style(style_kit, is_diff=False, fill=None, bold=None, horizontal=None):
    return dict(
        font=clone_font(
            style_kit["data_font"],
            bold=is_diff if bold is None else bold,
            italic=False,
            color=PALETTE["DIFF_CELL"] if is_diff else "FF000000",
        ),
        fill=copy(fill) if fill is not None else PatternFill(fill_type=None),
        alignment=clone_alignment(
            style_kit["data_alignment"],
            horizontal=horizontal if horizontal is not None else (style_kit["data_alignment"].horizontal or "left"),
            vertical=style_kit["data_alignment"].vertical or "center",
        ),
        border=copy(style_kit["data_border"]),
    )

def write_data_sheet(wb, title, df, table_name):
    ws = wb.create_sheet(title)
    if df.empty:
        ws.cell(1, 1, "No data").font = Font(italic=True)
        return

    cols = [c for c in COL_ORDER if c in df.columns]

    for ci, col in enumerate(cols, 1):
        ap(ws.cell(1, ci, col), hdr_style())
    ws.row_dimensions[1].height = 28

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        status = str(row.get("Status", ""))
        for ci, col in enumerate(cols, 1):
            val = row[col]
            if isinstance(val, float) and np.isnan(val):
                val = ""
            is_diff = (col in DIFF_COLS) and val not in ("", 0, 0.0) \
                      and not (isinstance(val, float) and abs(val) < 1e-9)
            c = ws.cell(ri, ci, val)
            ap(c, data_style(status, is_diff))
            if col not in ("Status", "ORIG_Name", "NEW_Name", "NAME_diff", "DIFF_Fields"):
                c.number_format = COORD_NUMBER_FORMAT
        ws.row_dimensions[ri].height = 18

    for ci, col in enumerate(cols, 1):
        sample = [str(col)] + [str(df.iloc[r, ci-1]) for r in range(min(300, len(df)))]
        w = min(max(len(v) + 2 for v in sample), 46)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ref = f"A1:{get_column_letter(len(cols))}{len(df)+1}"
    t = Table(displayName=table_name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)
    ws.freeze_panes = "B2"

def write_overview(wb, title, df_all, meta):
    ws = wb.create_sheet(title)

    for col, w in {"A": 3, "B": 32, "C": 14, "D": 10, "E": 42}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:E2")
    ws["B2"] = "Measurement Points Comparison Report"
    ws["B2"].font = Font(name=FONT_NAME, size=16, bold=True, color=PALETTE["TITLE"])
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:E3")
    ws["B3"] = (f"Workbook: {meta['workbook']}   |   "
                f"{datetime.now().strftime('%Y-%m-%d  %H:%M')}")
    ws["B3"].font = Font(name=FONT_NAME, size=10, italic=True, color="888888")
    ws.row_dimensions[3].height = 18

    ws.merge_cells("B4:E4")
    ws["B4"] = (f"Sheet 1: {meta['orig_sheet']} ({meta['n_orig']} rows)   |   "
                f"Sheet 2: {meta['new_sheet']} ({meta['n_new']} rows)   |   "
                f"XYZ tolerance: {meta['tol']} mm   |   "
                f"IJK: {'on' if meta['use_ijk'] else 'off'}")
    ws["B4"].font = Font(name=FONT_NAME, size=10, italic=True, color="AAAAAA")
    ws.row_dimensions[4].height = 16

    hrow = 6
    for ci, h in enumerate(["Status", "Count", "Color", "Description"], 2):
        ap(ws.cell(hrow, ci, h), hdr_style())
    ws.row_dimensions[hrow].height = 24

    vc = df_all["Status"].value_counts()
    for ri, code in enumerate(["MATCH","NAME_CHANGED","COORD_CHANGED","MOVED","DELETED","ADDED"], hrow+1):
        cnt = vc.get(code, 0)
        bg  = PALETTE.get(code, "FFFFFF")
        ws.row_dimensions[ri].height = 22
        for ci, val in enumerate([code, cnt, "", STATUS_LABEL[code]], 2):
            c = ws.cell(ri, ci, val)
            c.fill   = PatternFill("solid", fgColor=bg)
            c.font   = Font(name=FONT_NAME, size=11, bold=(ci in (2, 3)))
            c.border = thin()
            c.alignment = Alignment(
                horizontal="center" if ci in (2, 3, 4) else "left",
                vertical="center")

    tr = hrow + 6
    ws.row_dimensions[tr].height = 22
    for ci, val in enumerate(["TOTAL", len(df_all), "", "Total records"], 2):
        c = ws.cell(tr, ci, val)
        c.font   = Font(name=FONT_NAME, size=11, bold=True)
        c.border = thin()
        c.alignment = Alignment(
            horizontal="center" if ci in (2, 3, 4) else "left",
            vertical="center")

    lr = tr + 3
    ws.merge_cells(f"B{lr}:E{lr}")
    ws[f"B{lr}"] = "Orange text in diff columns means the value is not zero"
    ws[f"B{lr}"].font = Font(name=FONT_NAME, size=10, italic=True,
                              color=PALETTE["DIFF_CELL"])
    ws.row_dimensions[lr].height = 18

def delete_report_sheets(wb, prefix):
    for suffix in REPORT_SUFFIXES:
        name = f"{prefix}{suffix}"
        if name in wb.sheetnames:
            del wb[name]

def build_close_points_report(df_orig, df_new, distance_tol):
    """Find ALL original/new point pairs whose 3D XYZ distance is within distance_tol."""
    orig_xyz = df_orig[["X", "Y", "Z"]].to_numpy(dtype=float)
    new_xyz = df_new[["X", "Y", "Z"]].to_numpy(dtype=float)
    results = []

    for orig_idx, orig_row in df_orig.iterrows():
        orig_point = orig_xyz[orig_idx]
        if np.isnan(orig_point).any():
            continue

        deltas = new_xyz - orig_point
        axis_mask = ~np.isnan(deltas).any(axis=1)
        if not axis_mask.any():
            continue

        candidate_deltas = deltas[axis_mask]
        in_box = np.all(np.abs(candidate_deltas) <= distance_tol, axis=1)
        if not in_box.any():
            continue

        candidate_indices = np.where(axis_mask)[0][in_box]
        for new_idx, delta in zip(candidate_indices, candidate_deltas[in_box]):
            distance = float(np.linalg.norm(delta))
            if distance <= 1e-12 or distance > distance_tol + 1e-12:
                continue
            new_row = df_new.loc[new_idx]
            results.append({
                "ORIG_Name": orig_row["Name"],
                "ORIG_X": orig_row["X"],
                "ORIG_Y": orig_row["Y"],
                "ORIG_Z": orig_row["Z"],
                "NEW_Name": new_row["Name"],
                "NEW_X": new_row["X"],
                "NEW_Y": new_row["Y"],
                "NEW_Z": new_row["Z"],
                "dX": safe_delta(orig_row["X"], new_row["X"]),
                "dY": safe_delta(orig_row["Y"], new_row["Y"]),
                "dZ": safe_delta(orig_row["Z"], new_row["Z"]),
                "Distance_3D": round(distance, 4),
                "Same_Name": "Yes" if str(orig_row["Name"]) == str(new_row["Name"]) else "No",
            })

    if not results:
        return pd.DataFrame(columns=[
            "ORIG_Name", "ORIG_X", "ORIG_Y", "ORIG_Z",
            "NEW_Name", "NEW_X", "NEW_Y", "NEW_Z",
            "dX", "dY", "dZ", "Distance_3D", "Same_Name",
        ])

    return (
        pd.DataFrame(results)
        .sort_values(by=["Distance_3D", "ORIG_Name", "NEW_Name"], kind="stable")
        .reset_index(drop=True)
    )

def build_nearest_points_report(df_orig, df_new):
    """For each original point, find the SINGLE nearest new point (regardless of tolerance)."""
    orig_xyz = df_orig[["X", "Y", "Z"]].to_numpy(dtype=float)
    new_xyz = df_new[["X", "Y", "Z"]].to_numpy(dtype=float)
    results = []

    for orig_idx, orig_row in df_orig.iterrows():
        orig_point = orig_xyz[orig_idx]
        if np.isnan(orig_point).any():
            continue

        deltas = new_xyz - orig_point
        valid_mask = ~np.isnan(deltas).any(axis=1)
        if not valid_mask.any():
            continue

        valid_deltas = deltas[valid_mask]
        distances = np.linalg.norm(valid_deltas, axis=1)
        if len(distances) == 0:
            continue

        nearest_idx = int(np.argmin(distances))
        all_valid_indices = np.where(valid_mask)[0]
        new_idx = int(all_valid_indices[nearest_idx])
        distance = float(distances[nearest_idx])

        new_row = df_new.loc[new_idx]
        results.append({
            "ORIG_Name": orig_row["Name"],
            "ORIG_X": orig_row["X"],
            "ORIG_Y": orig_row["Y"],
            "ORIG_Z": orig_row["Z"],
            "NEW_Name": new_row["Name"],
            "NEW_X": new_row["X"],
            "NEW_Y": new_row["Y"],
            "NEW_Z": new_row["Z"],
            "dX": safe_delta(orig_row["X"], new_row["X"]),
            "dY": safe_delta(orig_row["Y"], new_row["Y"]),
            "dZ": safe_delta(orig_row["Z"], new_row["Z"]),
            "Distance_3D": round(distance, 4),
            "Same_Name": "Yes" if str(orig_row["Name"]) == str(new_row["Name"]) else "No",
        })

    if not results:
        return pd.DataFrame(columns=[
            "ORIG_Name", "ORIG_X", "ORIG_Y", "ORIG_Z",
            "NEW_Name", "NEW_X", "NEW_Y", "NEW_Z",
            "dX", "dY", "dZ", "Distance_3D", "Same_Name",
        ])

    return (
        pd.DataFrame(results)
        .sort_values(by=["Distance_3D", "ORIG_Name", "NEW_Name"], kind="stable")
        .reset_index(drop=True)
    )

def link_moved_pairs(df_all, df_orig, df_new, close_points_tol):
    """Pass 2: For DELETED originals + ADDED new points within close_points_tol, relabel both as MOVED.

    Returns (updated_df_all, moved_pairs_list).
    """
    deleted_mask = df_all["Status"] == "DELETED"
    added_mask = df_all["Status"] == "ADDED"

    df_deleted = df_all[deleted_mask].copy()
    df_added = df_all[added_mask].copy()

    if df_deleted.empty or df_added.empty:
        return df_all, []

    orig_xyz = df_orig[["X", "Y", "Z"]].to_numpy(dtype=float)
    new_xyz = df_new[["X", "Y", "Z"]].to_numpy(dtype=float)

    deleted_indices = list(df_deleted.index)
    added_indices = list(df_added.index)

    matched_deleted = set()
    matched_added = set()
    moved_pairs = []

    for del_idx in deleted_indices:
        orig_point = orig_xyz[del_idx]
        if np.isnan(orig_point).any():
            continue

        best_new_idx = None
        best_dist = float("inf")

        for add_idx in added_indices:
            if add_idx in matched_added:
                continue
            new_point = new_xyz[add_idx]
            if np.isnan(new_point).any():
                continue
            dist = float(np.linalg.norm(orig_point - new_point))
            if dist <= close_points_tol + 1e-12 and dist < best_dist:
                best_dist = dist
                best_new_idx = add_idx

        if best_new_idx is not None:
            matched_deleted.add(del_idx)
            matched_added.add(best_new_idx)
            orig_row = df_orig.loc[del_idx]
            new_row = df_new.loc[best_new_idx]
            moved_pairs.append({
                "ORIG_Name": orig_row["Name"],
                "ORIG_X": orig_row["X"],
                "ORIG_Y": orig_row["Y"],
                "ORIG_Z": orig_row["Z"],
                "NEW_Name": new_row["Name"],
                "NEW_X": new_row["X"],
                "NEW_Y": new_row["Y"],
                "NEW_Z": new_row["Z"],
                "dX": safe_delta(orig_row["X"], new_row["X"]),
                "dY": safe_delta(orig_row["Y"], new_row["Y"]),
                "dZ": safe_delta(orig_row["Z"], new_row["Z"]),
                "Distance_3D": round(best_dist, 4),
            })

    # Relabel matched rows in df_all
    df_all = df_all.copy()
    for idx in matched_deleted | matched_added:
        df_all.at[idx, "Status"] = "MOVED"

    return df_all, moved_pairs

def write_close_points_sheet(wb, title, df_close, meta):
    ws = wb.create_sheet(title)

    ws.merge_cells("A1:M1")
    ws["A1"] = f"Pairs of points within {meta['close_points_tol']:.3f} mm in 3D distance"
    ws["A1"].font = Font(name=FONT_NAME, size=14, bold=True, color=PALETTE["TITLE"])
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:M2")
    ws["A2"] = (
        f"Original sheet: {meta['orig_sheet']}   |   "
        f"New sheet: {meta['new_sheet']}   |   "
        f"Use this page to review possible same points that shifted slightly."
    )
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color="888888")
    ws.row_dimensions[2].height = 18

    if df_close.empty:
        ws.merge_cells("A4:M4")
        ws["A4"] = "No close point pairs were found within the XYZ tolerance radius."
        ws["A4"].font = Font(name=FONT_NAME, size=11, italic=True)
        return

    close_cols = list(df_close.columns)
    header_row = 4
    for ci, col in enumerate(close_cols, 1):
        ap(ws.cell(header_row, ci, col), hdr_style())
    ws.row_dimensions[header_row].height = 24

    for ri, (_, row) in enumerate(df_close.iterrows(), header_row + 1):
        for ci, col in enumerate(close_cols, 1):
            value = row[col]
            if isinstance(value, float) and np.isnan(value):
                value = ""
            cell = ws.cell(ri, ci, value)
            ap(cell, data_style("ADDED"))
            if col.endswith(("_X", "_Y", "_Z")) or col in {"dX", "dY", "dZ", "Distance_3D"}:
                cell.number_format = COORD_NUMBER_FORMAT
        ws.row_dimensions[ri].height = 18

    for ci, col in enumerate(close_cols, 1):
        sample = [str(col)] + [str(df_close.iloc[r, ci - 1]) for r in range(min(300, len(df_close)))]
        width = min(max(len(v) + 2 for v in sample), 24)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ref = f"A{header_row}:{get_column_letter(len(close_cols))}{len(df_close) + header_row}"
    table = Table(displayName="ClosePoints", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = f"B{header_row + 1}"

def write_nearest_points_sheet(wb, title, df_nearest, meta):
    ws = wb.create_sheet(title)

    ws.merge_cells("A1:M1")
    ws["A1"] = "Nearest new point for each original point"
    ws["A1"].font = Font(name=FONT_NAME, size=14, bold=True, color=PALETTE["TITLE"])
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:M2")
    ws["A2"] = (
        f"Original sheet: {meta['orig_sheet']}   |   "
        f"New sheet: {meta['new_sheet']}   |   "
        f"Each original point is paired with its closest new point regardless of distance."
    )
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color="888888")
    ws.row_dimensions[2].height = 18

    if df_nearest.empty:
        ws.merge_cells("A4:M4")
        ws["A4"] = "No nearest-point pairs could be computed (missing coordinates)."
        ws["A4"].font = Font(name=FONT_NAME, size=11, italic=True)
        return

    nearest_cols = list(df_nearest.columns)
    header_row = 4
    for ci, col in enumerate(nearest_cols, 1):
        ap(ws.cell(header_row, ci, col), hdr_style())
    ws.row_dimensions[header_row].height = 24

    for ri, (_, row) in enumerate(df_nearest.iterrows(), header_row + 1):
        for ci, col in enumerate(nearest_cols, 1):
            value = row[col]
            if isinstance(value, float) and np.isnan(value):
                value = ""
            cell = ws.cell(ri, ci, value)
            ap(cell, data_style("ADDED"))
            if col.endswith(("_X", "_Y", "_Z")) or col in {"dX", "dY", "dZ", "Distance_3D"}:
                cell.number_format = COORD_NUMBER_FORMAT
        ws.row_dimensions[ri].height = 18

    for ci, col in enumerate(nearest_cols, 1):
        sample = [str(col)] + [str(df_nearest.iloc[r, ci - 1]) for r in range(min(300, len(df_nearest)))]
        width = min(max(len(v) + 2 for v in sample), 24)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ref = f"A{header_row}:{get_column_letter(len(nearest_cols))}{len(df_nearest) + header_row}"
    table = Table(displayName="NearestPoints", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = f"B{header_row + 1}"

def write_report_sheets(wb, df_all, meta, prefix=REPORT_PREFIX):
    delete_report_sheets(wb, prefix)
    df_close = meta["df_close"]
    write_overview(wb, f"{prefix}Overview", df_all, meta)
    write_data_sheet(wb, f"{prefix}All Results",  df_all,                                    "AllResults")
    write_data_sheet(wb, f"{prefix}Match",        df_all[df_all.Status=="MATCH"].reset_index(drop=True),         "Match")
    write_data_sheet(wb, f"{prefix}Name Changed", df_all[df_all.Status=="NAME_CHANGED"].reset_index(drop=True),  "NameChanged")
    write_data_sheet(wb, f"{prefix}Coord Changed", df_all[df_all.Status=="COORD_CHANGED"].reset_index(drop=True), "CoordChanged")
    write_data_sheet(wb, f"{prefix}Moved",        df_all[df_all.Status=="MOVED"].reset_index(drop=True),         "Moved")
    write_data_sheet(wb, f"{prefix}Deleted",      df_all[df_all.Status=="DELETED"].reset_index(drop=True),       "Deleted")
    write_data_sheet(wb, f"{prefix}Added",        df_all[df_all.Status=="ADDED"].reset_index(drop=True),         "Added")
    write_close_points_sheet(wb, f"{prefix}Close Points", df_close, meta)

def sanitize_external_query_metadata(wb):
    """Convert external/query-style tables into regular worksheet tables before save.

    openpyxl does not preserve Excel's full query/externalData package parts. If a
    workbook contains imported query tables, saving it as-is leaves dangling
    query-range metadata and Excel repairs the file on open. We normalize those
    tables into regular worksheet tables and remove their hidden ExternalData_*
    defined names so the workbook opens cleanly.
    """
    for ws in wb.worksheets:
        for table in ws._tables.values():
            if getattr(table, "tableType", None) == "queryTable":
                table.tableType = "worksheet"
                for column in table.tableColumns:
                    if hasattr(column, "queryTableFieldId"):
                        column.queryTableFieldId = None

EXCEL_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

def clean_workbook_xml_bytes(data):
    root = ET.fromstring(data)
    defined_names = root.find(f"{{{EXCEL_MAIN_NS}}}definedNames")
    if defined_names is None:
        return data

    for defined_name in list(defined_names):
        if defined_name.attrib.get("name", "").startswith("ExternalData_"):
            defined_names.remove(defined_name)

    if not list(defined_names):
        root.remove(defined_names)

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)

def clean_table_xml_bytes(data):
    root = ET.fromstring(data)
    if root.attrib.get("tableType") == "queryTable":
        root.attrib["tableType"] = "worksheet"

    for column in root.findall(f".//{{{EXCEL_MAIN_NS}}}tableColumn"):
        column.attrib.pop("queryTableFieldId", None)

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)

def clean_saved_workbook_package(workbook_path):
    workbook_path = Path(workbook_path)
    with NamedTemporaryFile(delete=False, suffix=workbook_path.suffix) as tmp:
        temp_path = Path(tmp.name)

    with ZipFile(workbook_path, "r") as src_zip, ZipFile(temp_path, "w") as dst_zip:
        for info in src_zip.infolist():
            data = src_zip.read(info.filename)
            if info.filename == "xl/workbook.xml":
                data = clean_workbook_xml_bytes(data)
            elif info.filename.startswith("xl/tables/") and info.filename.endswith(".xml"):
                data = clean_table_xml_bytes(data)
            dst_zip.writestr(info, data)

    # Use shutil.move instead of Path.replace so the operation works even when
    # the temp directory and the workbook reside on different drives (WinError 17).
    shutil.move(str(temp_path), str(workbook_path))

def save_workbook_clean(wb, workbook_path):
    sanitize_external_query_metadata(wb)
    wb.save(workbook_path)
    clean_saved_workbook_package(workbook_path)

def build_report(df_all, meta, workbook_path, prefix=REPORT_PREFIX):
    keep_vba = Path(workbook_path).suffix.lower() == ".xlsm"
    wb = load_workbook(workbook_path, keep_vba=keep_vba)
    write_report_sheets(wb, df_all, meta, prefix=prefix)
    save_workbook_clean(wb, workbook_path)

def find_existing_append_block(ws, headers, legacy_headers=None, header_row=None):
    row_values = []
    if header_row is None:
        header_row = get_original_sheet_header_row(ws)
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_idx).value
        row_values.append("" if value is None else str(value).strip())

    header_sets = [headers]
    if legacy_headers:
        header_sets.append(legacy_headers)

    for header_set in header_sets:
        block_len = len(header_set)
        for start_idx in range(0, len(row_values) - block_len + 1):
            if row_values[start_idx:start_idx + block_len] == header_set:
                return start_idx + 1
    return None

def get_original_sheet_header_row(ws):
    row1 = ["" if ws.cell(1, col_idx).value is None else str(ws.cell(1, col_idx).value).strip() for col_idx in range(1, ws.max_column + 1)]
    row2 = ["" if ws.cell(2, col_idx).value is None else str(ws.cell(2, col_idx).value).strip() for col_idx in range(1, ws.max_column + 1)]
    if any(value in GROUP_HEADER_LABELS for value in row1) and all(field in row2 for field in LOGICAL_FIELDS):
        return 2
    return 1

def ensure_original_sheet_group_row(ws):
    header_row = get_original_sheet_header_row(ws)
    if header_row == 1:
        ws.insert_rows(1)
        header_row = 2
    return header_row

def disable_sheet_filter(ws):
    if getattr(ws.auto_filter, "ref", None):
        ws.auto_filter.ref = None

    for row_idx in range(1, ws.max_row + 1):
        if row_idx in ws.row_dimensions:
            ws.row_dimensions[row_idx].hidden = False

def format_original_coordinate_columns(ws, header_row, data_start_row, source_end_col):
    source_coord_columns = {}
    for col_idx in range(1, source_end_col + 1):
        header_value = ws.cell(header_row, col_idx).value
        if header_value is None:
            continue
        normalized = str(header_value).strip()
        if normalized in COORD_FIELDS and normalized not in source_coord_columns:
            source_coord_columns[normalized] = col_idx

    for col_idx in source_coord_columns.values():
        for row_idx in range(data_start_row, ws.max_row + 1):
            ws.cell(row_idx, col_idx).number_format = COORD_NUMBER_FORMAT

def clear_existing_output_rows(ws, data_start_row, original_data_rows, source_end_col):
    row_idx = data_start_row + original_data_rows
    while row_idx <= ws.max_row:
        has_original_data = any(
            ws.cell(row_idx, col_idx).value not in (None, "")
            for col_idx in range(1, source_end_col + 1)
        )
        if has_original_data:
            break
        ws.delete_rows(row_idx, 1)

def apply_group_labels(ws, style_kit, source_end_col, status_col, new_start_col, new_end_col, diff_start_col, diff_end_col):
    for merge_range in list(ws.merged_cells.ranges):
        if merge_range.min_row == 1 and merge_range.max_row == 1:
            ws.unmerge_cells(str(merge_range))

    def write_group(start_col, end_col, label):
        if start_col > end_col:
            return
        if start_col == end_col:
            cell = ws.cell(1, start_col, label)
            ap(cell, original_sheet_group_style(style_kit, label))
            return
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(1, start_col, label)
        ap(cell, original_sheet_group_style(style_kit, label))

    write_group(1, source_end_col, "Original data")
    cell = ws.cell(1, status_col, STATUS_GROUP_LABEL)
    ap(cell, original_sheet_group_style(style_kit, STATUS_GROUP_LABEL))
    write_group(new_start_col, new_end_col, "New Data")
    write_group(diff_start_col, diff_end_col, "Difference")

def append_results_to_original_sheet(wb, sheet_name, df_all, n_orig):
    ws = wb[sheet_name]
    disable_sheet_filter(ws)
    header_row = ensure_original_sheet_group_row(ws)
    headers = [header for _, header in ORIGINAL_SHEET_APPEND_MAP]
    start_col = find_existing_append_block(
        ws,
        headers,
        legacy_headers=LEGACY_ORIGINAL_SHEET_APPEND_HEADERS,
        header_row=header_row,
    )
    source_end_col = start_col - 1 if start_col is not None else ws.max_column
    if start_col is None:
        start_col = source_end_col + 1
    end_col = start_col + len(headers) - 1
    data_start_row = header_row + 1
    style_kit = infer_original_sheet_style_kit(ws, header_row, data_start_row, source_end_col)
    format_original_coordinate_columns(ws, header_row, data_start_row, source_end_col)

    for col_idx in range(start_col, end_col + 1):
        for row_idx in range(header_row, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.alignment = Alignment(horizontal="general", vertical="bottom")
            cell.font = Font(name=FONT_NAME, size=10)

    clear_existing_output_rows(ws, data_start_row, n_orig, source_end_col)

    for offset, (_, header) in enumerate(ORIGINAL_SHEET_APPEND_MAP):
        cell = ws.cell(header_row, start_col + offset, header)
        fill = PatternFill(patternType="solid", fgColor="FFEDEDED") if header == "STATUS" else None
        ap(cell, original_sheet_header_style(style_kit, fill=fill))

    orig_rows = df_all.iloc[:n_orig].reset_index(drop=True)
    for row_idx, (_, result_row) in enumerate(orig_rows.iterrows(), start=data_start_row):
        for offset, (source_col, _) in enumerate(ORIGINAL_SHEET_APPEND_MAP):
            value = result_row[source_col] if source_col in result_row else ""
            if isinstance(value, float) and np.isnan(value):
                value = ""
            is_diff = source_col.endswith("_diff") or source_col == "DIFF_Fields"
            cell = ws.cell(row_idx, start_col + offset, value)
            if source_col == "Status":
                ap(
                    cell,
                    original_sheet_data_style(
                        style_kit,
                        is_diff=False,
                        fill=STATUS_FILLS.get(str(value), copy(GROUP_FILL_ORIGINAL)),
                        bold=True,
                        horizontal="center",
                    ),
                )
            else:
                ap(cell, original_sheet_data_style(style_kit, is_diff=is_diff and value not in ("", 0, 0.0)))
            if source_col in ORIGINAL_SHEET_NUMERIC_COLUMNS:
                cell.number_format = COORD_NUMBER_FORMAT

    added_rows = df_all[df_all.Status == "ADDED"].reset_index(drop=True)
    append_row_idx = data_start_row + n_orig
    for _, added_row in added_rows.iterrows():
        for offset, (source_col, _) in enumerate(ORIGINAL_SHEET_APPEND_MAP):
            value = ""
            if source_col == "Status":
                value = "NEW"
            elif source_col.startswith("NEW_"):
                value = added_row.get(source_col, "")
                if isinstance(value, float) and np.isnan(value):
                    value = ""
            cell = ws.cell(append_row_idx, start_col + offset, value)
            is_diff = source_col.endswith("_diff") or source_col == "DIFF_Fields"
            if source_col == "Status":
                ap(
                    cell,
                    original_sheet_data_style(
                        style_kit,
                        is_diff=False,
                        fill=STATUS_FILLS["NEW"],
                        bold=True,
                        horizontal="center",
                    ),
                )
            else:
                ap(cell, original_sheet_data_style(style_kit, is_diff=is_diff and value not in ("", 0, 0.0)))
            if source_col in ORIGINAL_SHEET_NUMERIC_COLUMNS:
                cell.number_format = COORD_NUMBER_FORMAT
        append_row_idx += 1

    for offset, (_, header) in enumerate(ORIGINAL_SHEET_APPEND_MAP):
        col_idx = start_col + offset
        sample_values = [str(header)]
        sample_values.extend(
            str(ws.cell(row_idx, col_idx).value)
            for row_idx in range(data_start_row, ws.max_row + 1)
            if ws.cell(row_idx, col_idx).value not in (None, "")
        )
        width = min(max((len(value) + 2) for value in sample_values), 24)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    apply_group_labels(
        ws,
        style_kit,
        source_end_col=source_end_col,
        status_col=start_col,
        new_start_col=start_col + 1,
        new_end_col=start_col + 7,
        diff_start_col=start_col + 8,
        diff_end_col=end_col,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  SHARED EXECUTION FLOW
# ══════════════════════════════════════════════════════════════════════════════

def run_comparison(
    workbook_path,
    orig_sheet,
    new_sheet,
    tol,
    ijk_tol,
    use_ijk,
    create_report_sheets=True,
    close_points_tol=None,
):
    """Run the workbook comparison and write report sheets back into the workbook."""
    if close_points_tol is None:
        close_points_tol = DEFAULTS["close_points_tol"]

    df_orig_raw = read_sheet(workbook_path, orig_sheet)
    df_new_raw = read_sheet(workbook_path, new_sheet)

    df_orig = prepare_sheet(df_orig_raw, orig_sheet)
    df_new = prepare_sheet(df_new_raw, new_sheet)

    # Pass 1: Standard comparison
    df_all = compare(df_orig, df_new, tol, ijk_tol, use_ijk)

    # Pass 2: Link DELETED+ADDED pairs within close_points_tol → MOVED
    df_all, moved_pairs = link_moved_pairs(df_all, df_orig, df_new, close_points_tol)

    # Build close-points report (includes MOVED pairs + any other close pairs)
    df_close = build_close_points_report(df_orig, df_new, close_points_tol)

    meta = {
        "workbook": Path(workbook_path).name,
        "orig_sheet": orig_sheet,
        "new_sheet": new_sheet,
        "n_orig": len(df_orig),
        "n_new": len(df_new),
        "tol": tol,
        "ijk_tol": ijk_tol,
        "use_ijk": use_ijk,
        "df_orig": df_orig,
        "df_new": df_new,
        "df_close": df_close,
        "moved_pairs": moved_pairs,
        "close_points_tol": close_points_tol,
    }
    keep_vba = Path(workbook_path).suffix.lower() == ".xlsm"
    wb = load_workbook(workbook_path, keep_vba=keep_vba)
    append_results_to_original_sheet(wb, orig_sheet, df_all, len(df_orig))
    if create_report_sheets:
        write_report_sheets(wb, df_all, meta)
    save_workbook_clean(wb, workbook_path)
    return {
        "df_all": df_all,
        "meta": meta,
        "orig_columns": list(df_orig_raw.columns),
        "new_columns": list(df_new_raw.columns),
        "create_report_sheets": create_report_sheets,
        "close_points_count": len(df_close),
        "moved_count": len(moved_pairs),
    }


# ══════════════════════════════════════════════════════════════════════════════
#  INTERACTIVE FLOW
# ══════════════════════════════════════════════════════════════════════════════

def interactive():
    hr("═")
    print("  MEASUREMENT POINTS COMPARISON")
    hr("═")
    print()

    hr()
    print("  STEP 1 OF 4 - EXCEL WORKBOOK")
    hr()
    workbook_path = ask("  Path to the Excel workbook", validator=validate_workbook)

    sheets = list_sheets(workbook_path)
    if not sheets:
        print("\n  ⚠  Unable to read worksheet names from the Excel file.")
        return None, None

    print(f"\n  Available sheets: {', '.join(str(s) for s in sheets)}")

    print()
    hr()
    print("  STEP 2 OF 4 - ORIGINAL SHEET")
    hr()
    orig_sheet = ask("  Name of the first sheet", default=str(sheets[0]))
    orig_sheet = orig_sheet if orig_sheet in sheets else sheets[0]

    print()
    hr()
    print("  STEP 3 OF 4 - NEW SHEET")
    hr()
    default_new_sheet = str(sheets[1]) if len(sheets) > 1 else str(sheets[0])
    new_sheet = ask("  Name of the second sheet", default=default_new_sheet)
    new_sheet = new_sheet if new_sheet in sheets else default_new_sheet

    print()
    hr()
    print("  Reading Excel sheets ...")
    hr()
    df_orig_raw = read_sheet(workbook_path, orig_sheet)
    df_new_raw = read_sheet(workbook_path, new_sheet)
    print(f"  Sheet 1: {orig_sheet}  |  Rows: {len(df_orig_raw)}  |  Columns: {list(df_orig_raw.columns)}")
    print(f"  Sheet 2: {new_sheet}  |  Rows: {len(df_new_raw)}  |  Columns: {list(df_new_raw.columns)}")

    print()
    hr()
    print("  STEP 4 OF 4 - COMPARISON SETTINGS")
    hr()
    tol     = ask_float("  XYZ tolerance (mm)", DEFAULTS["coord_tol"])
    ijk_tol = ask_float("  IJK tolerance",       DEFAULTS["ijk_tol"])
    use_ijk = ask_yn("  Include I/J/K in coordinate comparison?", DEFAULTS["compare_ijk"])
    close_tol = ask_float("  Close-points 3D tolerance (mm)", DEFAULTS["close_points_tol"])

    print()
    hr()
    print("  Normalizing data ...")
    try:
        result = run_comparison(workbook_path, orig_sheet, new_sheet, tol, ijk_tol, use_ijk, close_points_tol=close_tol)
    except ValueError as exc:
        print(f"  ⚠  {exc}")
        hr("═")
        return None, None

    df_all = result["df_all"]
    meta = result["meta"]

    vc = df_all["Status"].value_counts()
    print()
    hr()
    print("  RESULTS")
    hr()
    for code in ["MATCH","NAME_CHANGED","COORD_CHANGED","MOVED","DELETED","ADDED"]:
        print(f"  {STATUS_LABEL[code]:<40s}: {vc.get(code, 0):>5d}")
    print(f"  {'TOTAL':<40s}: {len(df_all):>5d}")
    print(f"  {'Close 3D pairs (within tolerance)':<40s}: {result['close_points_count']:>5d}")
    print(f"  {'Moved pairs (DELETED+ADDED within tolerance)':<40s}: {result['moved_count']:>5d}")
    hr()

    print(f"\n  ✅ Done. The original sheet was updated and CMP_* sheets were written to {Path(workbook_path).resolve()}")
    hr("═")

    return df_all, workbook_path


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    try:
        interactive()
    except KeyboardInterrupt:
        print("\n\n  Cancelled by user.")
        sys.exit(0)
