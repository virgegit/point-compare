"""
Schema and constants for point-compare comparison tool.

Defines field names, tolerances, color palettes, status labels, and other constants
shared across all implementations (interactive CLI, batch, GUI).
"""

from openpyxl.styles import PatternFill, Color
from copy import copy

# ══════════════════════════════════════════════════════════════════════════════
#  DEFAULT SETTINGS
# ══════════════════════════════════════════════════════════════════════════════
DEFAULTS = {
    "coord_tol":       0.05,   # XYZ tolerance, mm
    "ijk_tol":         0.001,  # IJK tolerance
    "compare_ijk":     False,  # include I/J/K in coordinate comparison?
    "close_points_tol": 10.0,  # 3D distance threshold for close-points report, mm
}

# ══════════════════════════════════════════════════════════════════════════════
#  FIELD DEFINITIONS
# ══════════════════════════════════════════════════════════════════════════════
LOGICAL_FIELDS = ["Name", "X", "Y", "Z", "I", "J", "K"]
COORD_FIELDS = ["X", "Y", "Z", "I", "J", "K"]

# ══════════════════════════════════════════════════════════════════════════════
#  STYLE PALETTE
# ══════════════════════════════════════════════════════════════════════════════
PALETTE = {
    "MATCH":         "C6EFCE",
    "NAME_CHANGED":  "FFEB9C",
    "COORD_CHANGED": "FFEB9C",
    "REPLACED?":     "FFEB9C",
    "DELETED":       "F4CCCC",
    "ADDED":         "D9EAD3",
    "DIFF_CELL":     "CC3300",
    "HEADER_BG":     "4472C4",
    "HEADER_FG":     "FFFFFF",
    "TITLE":         "1F3864",
}

STATUS_LABEL = {
    "MATCH":         "✔  Match",
    "NAME_CHANGED":  "✎  Name Changed",
    "COORD_CHANGED": "⚠  Coordinates Changed",
    "REPLACED?":     "↔  Replaced?",
    "DELETED":       "✖  Deleted (only in original)",
    "ADDED":         "＋ Added (only in new)",
}

# ══════════════════════════════════════════════════════════════════════════════
#  REPORT CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
REPORT_PREFIX = "CMP_"
REPORT_SUFFIXES = [
    "Overview",
    "All Results",
    "Match",
    "Name Changed",
    "Coord Changed",
    "Replaced",
    "Deleted",
    "Added",
    "Close Points",
]

# ══════════════════════════════════════════════════════════════════════════════
#  FORMATTING
# ══════════════════════════════════════════════════════════════════════════════
FONT_NAME = "Calibri"
COORD_NUMBER_FORMAT = "0.000"

# ══════════════════════════════════════════════════════════════════════════════
#  ORIGINAL SHEET LAYOUT
# ══════════════════════════════════════════════════════════════════════════════
ORIGINAL_SHEET_APPEND_MAP = [
    ("Status", "STATUS"),
    ("NEW_Name", "Name"),
    ("NEW_X", "X"),
    ("NEW_Y", "Y"),
    ("NEW_Z", "Z"),
    ("NEW_I", "I"),
    ("NEW_J", "J"),
    ("NEW_K", "K"),
    ("NAME_diff", "NAME_diff"),
    ("X_diff", "X_diff"),
    ("Y_diff", "Y_diff"),
    ("Z_diff", "Z_diff"),
    ("I_diff", "I_diff"),
    ("J_diff", "J_diff"),
    ("K_diff", "K_diff"),
    ("DIFF_Fields", "DIFF_Fields"),
]

LEGACY_ORIGINAL_SHEET_APPEND_HEADERS = [
    "STATUS",
    "NEW_Name",
    "NEW_X",
    "NEW_Y",
    "NEW_Z",
    "NEW_I",
    "NEW_J",
    "NEW_K",
    "NAME_diff",
    "X_diff",
    "Y_diff",
    "Z_diff",
    "I_diff",
    "J_diff",
    "K_diff",
    "DIFF_Fields",
]

ORIGINAL_SHEET_NUMERIC_COLUMNS = {
    "NEW_X", "NEW_Y", "NEW_Z", "NEW_I", "NEW_J", "NEW_K",
    "X_diff", "Y_diff", "Z_diff", "I_diff", "J_diff", "K_diff",
}

# ══════════════════════════════════════════════════════════════════════════════
#  GROUP HEADERS (for original sheet row 1)
# ══════════════════════════════════════════════════════════════════════════════
GROUP_HEADER_LABELS = {"Original data", "New Data", "Difference", "STATUS", "Comparison \nstatus"}
STATUS_GROUP_LABEL = "Comparison \nstatus"

# ══════════════════════════════════════════════════════════════════════════════
#  GROUP FILL COLORS
# ══════════════════════════════════════════════════════════════════════════════
LIGHT_THEME_COLOR = Color(theme=9, tint=0.7999816888943144)
GROUP_FILL_ORIGINAL = PatternFill(patternType="solid", fgColor=copy(LIGHT_THEME_COLOR))
GROUP_FILL_STATUS = PatternFill(patternType="solid", fgColor="FF5B9BD5")
GROUP_FILL_NEW = PatternFill(patternType="solid", fgColor="FFFFFF00")
GROUP_FILL_DIFF = PatternFill(patternType="solid", fgColor="FF5B9BD5")

STATUS_FILLS = {
    "MATCH": PatternFill(patternType="solid", fgColor=copy(LIGHT_THEME_COLOR)),
    "NAME_CHANGED": PatternFill(patternType="solid", fgColor="FFFFFF00"),
    "REPLACED?": PatternFill(patternType="solid", fgColor="FFFFFF00"),
    "COORD_CHANGED": PatternFill(patternType="solid", fgColor="FFFFFF00"),
    "DELETED": PatternFill(patternType="solid", fgColor="FFFF0000"),
    "NEW": PatternFill(patternType="solid", fgColor="FFFFC000"),
}

# ══════════════════════════════════════════════════════════════════════════════
#  REPORT COLUMN ORDERING
# ══════════════════════════════════════════════════════════════════════════════
COL_ORDER = [
    "Status",
    "ORIG_Name","ORIG_X","ORIG_Y","ORIG_Z","ORIG_I","ORIG_J","ORIG_K",
    "NEW_Name", "NEW_X", "NEW_Y", "NEW_Z", "NEW_I", "NEW_J", "NEW_K",
    "NAME_diff","X_diff","Y_diff","Z_diff","I_diff","J_diff","K_diff",
    "DIFF_Fields",
]

DIFF_COLS = {"NAME_diff","X_diff","Y_diff","Z_diff","I_diff","J_diff","K_diff","DIFF_Fields"}
