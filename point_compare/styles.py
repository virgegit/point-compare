"""Styling utilities and utilities for point-compare."""

from copy import copy
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .schema import PALETTE, FONT_NAME, GROUP_FILL_ORIGINAL, GROUP_FILL_STATUS, GROUP_FILL_NEW, GROUP_FILL_DIFF, STATUS_FILLS


def thin(color="CCCCCC"):
    """Create thin border with specified color."""
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def ap(cell, style):
    """Apply style dictionary to openpyxl cell."""
    for k, v in style.items():
        setattr(cell, k, v)


def hdr_style():
    """Style for table headers."""
    return dict(
        font=Font(name=FONT_NAME, bold=True, color=PALETTE["HEADER_FG"], size=11),
        fill=PatternFill("solid", fgColor=PALETTE["HEADER_BG"]),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=thin(),
    )


def data_style(status, is_diff=False):
    """Style for data cells based on status."""
    bg = PALETTE.get(status, "FFFFFF")
    if is_diff:
        font = Font(name=FONT_NAME, size=10, bold=True, color=PALETTE["DIFF_CELL"])
    else:
        font = Font(name=FONT_NAME, size=10)
    return dict(
        font=font,
        fill=PatternFill("solid", fgColor=bg),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=thin(),
    )


def clone_font(base_font, **updates):
    """Clone font with updated properties."""
    new_font = copy(base_font)
    for key, value in updates.items():
        setattr(new_font, key, value)
    return new_font


def clone_alignment(base_alignment, **updates):
    """Clone alignment with updated properties."""
    new_alignment = copy(base_alignment)
    for key, value in updates.items():
        setattr(new_alignment, key, value)
    return new_alignment


def copy_style_or_default(value, default_factory):
    """Copy style or use default factory if None."""
    return copy(value) if value is not None else default_factory()


def find_first_non_empty_in_row(ws, row_idx, start_col, end_col):
    """Find first non-empty cell in row range."""
    for col_idx in range(start_col, end_col + 1):
        if ws.cell(row_idx, col_idx).value not in (None, ""):
            return ws.cell(row_idx, col_idx)
    return None


def find_first_non_empty_data_cell(ws, data_start_row, source_end_col):
    """Find first non-empty cell in data range."""
    for row_idx in range(data_start_row, ws.max_row + 1):
        for col_idx in range(1, source_end_col + 1):
            if ws.cell(row_idx, col_idx).value not in (None, ""):
                return ws.cell(row_idx, col_idx)
    return None


def infer_original_sheet_style_kit(ws, header_row, data_start_row, source_end_col):
    """Infer styles from existing original sheet."""
    header_sample = find_first_non_empty_in_row(ws, header_row, 1, source_end_col)
    data_sample = find_first_non_empty_data_cell(ws, data_start_row, source_end_col)

    if header_sample is None:
        header_sample = ws.cell(header_row, 1)
    if data_sample is None:
        data_sample = ws.cell(data_start_row, 1)

    return {
        "header_font": copy_style_or_default(header_sample.font, lambda: Font(name=FONT_NAME, size=11, bold=True)),
        "header_fill": copy_style_or_default(header_sample.fill, lambda: PatternFill(patternType="solid", fgColor="FFFFFF")),
        "header_alignment": copy_style_or_default(header_sample.alignment, lambda: Alignment(horizontal="center", vertical="center")),
        "header_border": copy_style_or_default(header_sample.border, thin),
        "data_font": copy_style_or_default(data_sample.font, lambda: Font(name=FONT_NAME, size=11)),
        "data_alignment": copy_style_or_default(data_sample.alignment, lambda: Alignment(horizontal="left", vertical="center")),
        "data_border": copy_style_or_default(data_sample.border, thin),
    }


def original_sheet_header_style(style_kit, fill=None):
    """Style for original sheet headers."""
    return dict(
        font=clone_font(style_kit["header_font"], color="FF222222", bold=True, italic=False),
        fill=copy(fill) if fill is not None else PatternFill(patternType="solid", fgColor="FFEDEDED"),
        alignment=clone_alignment(style_kit["header_alignment"], horizontal="center", vertical="center", wrap_text=True),
        border=copy(style_kit["header_border"]),
    )


def original_sheet_group_style(style_kit, label):
    """Style for original sheet group labels."""
    from .schema import STATUS_GROUP_LABEL
    if label == "Original data":
        fill, size = GROUP_FILL_ORIGINAL, 16
    elif label == STATUS_GROUP_LABEL:
        fill, size = GROUP_FILL_STATUS, 11
    elif label == "New Data":
        fill, size = GROUP_FILL_NEW, 16
    else:
        fill, size = GROUP_FILL_DIFF, 16

    return dict(
        font=clone_font(style_kit["header_font"], size=size, bold=True, italic=False),
        fill=copy(fill),
        alignment=clone_alignment(style_kit["header_alignment"], horizontal="center", vertical="center", wrap_text=True),
        border=copy(style_kit["header_border"]),
    )


def original_sheet_data_style(style_kit, is_diff=False, fill=None, bold=None, horizontal=None):
    """Style for original sheet data cells."""
    return dict(
        font=clone_font(
            style_kit["data_font"],
            bold=is_diff if bold is None else bold,
            italic=False,
            color=PALETTE["DIFF_CELL"] if is_diff else "FF000000",
        ),
        fill=copy(fill) if fill is not None else PatternFill(fill_type=None),
        alignment=clone_alignment(
            style_kit["data_alignment"],
            horizontal=horizontal if horizontal is not None else (style_kit["data_alignment"].horizontal or "left"),
            vertical=style_kit["data_alignment"].vertical or "center",
        ),
        border=copy(style_kit["data_border"]),
    )
