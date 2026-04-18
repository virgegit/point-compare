"""Excel I/O operations for point-compare."""

import sys
import shutil
from pathlib import Path
from datetime import datetime
from tempfile import NamedTemporaryFile
from xml.etree import ElementTree as ET
from zipfile import ZipFile

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .schema import (
    DEFAULTS, LOGICAL_FIELDS, COORD_FIELDS, COORD_NUMBER_FORMAT,
    FONT_NAME, REPORT_PREFIX, REPORT_SUFFIXES, GROUP_HEADER_LABELS,
    ORIGINAL_SHEET_APPEND_MAP, LEGACY_ORIGINAL_SHEET_APPEND_HEADERS,
    ORIGINAL_SHEET_NUMERIC_COLUMNS, COL_ORDER, DIFF_COLS, PALETTE,
    STATUS_LABEL, STATUS_FILLS, GROUP_FILL_ORIGINAL
)
from .core import compare, build_close_points_report, build_nearest_points_report, link_moved_pairs
from .styles import thin, ap, hdr_style, data_style, infer_original_sheet_style_kit, original_sheet_header_style, original_sheet_group_style, original_sheet_data_style


def detect_header_row(path, sheet):
    """Detect if data starts on row 1 or row 2 (with group labels)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        row1 = ["" if cell.value is None else str(cell.value).strip() for cell in ws[1]]
        row2 = ["" if cell.value is None else str(cell.value).strip() for cell in ws[2]]
    finally:
        wb.close()
    if any(value in GROUP_HEADER_LABELS for value in row1) and all(field in row2 for field in LOGICAL_FIELDS):
        return 2
    return 1


def read_sheet(path, sheet):
    """Read an Excel sheet into a DataFrame."""
    header_row = detect_header_row(path, sheet)
    return pd.read_excel(path, sheet_name=sheet, header=header_row - 1)


def prepare_sheet(df, sheet_name):
    """Validate required columns and normalize the sheet."""
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    missing = [field for field in LOGICAL_FIELDS if field not in df.columns]
    if missing:
        raise ValueError(f"Required columns missing on sheet '{sheet_name}': {', '.join(missing)}")

    out = {}
    for field in LOGICAL_FIELDS:
        if field == "Name":
            out[field] = df[field].where(df[field].notna(), "").astype(str).str.strip()
        else:
            out[field] = pd.to_numeric(df[field], errors="coerce")
    prepared = pd.DataFrame(out)
    keep_mask = (prepared["Name"] != "") | prepared[["X", "Y", "Z", "I", "J", "K"]].notna().any(axis=1)
    return prepared.loc[keep_mask].reset_index(drop=True)


def write_data_sheet(wb, title, df, table_name):
    """Write comparison results to a new sheet."""
    ws = wb.create_sheet(title)
    if df.empty:
        ws.cell(1, 1, "No data").font = Font(italic=True)
        return

    cols = [c for c in COL_ORDER if c in df.columns]
    for ci, col in enumerate(cols, 1):
        ap(ws.cell(1, ci, col), hdr_style())
    ws.row_dimensions[1].height = 28

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        status = str(row.get("Status", ""))
        for ci, col in enumerate(cols, 1):
            val = row[col]
            if isinstance(val, float) and np.isnan(val):
                val = ""
            is_diff = (col in DIFF_COLS) and val not in ("", 0, 0.0) and not (isinstance(val, float) and abs(val) < 1e-9)
            c = ws.cell(ri, ci, val)
            ap(c, data_style(status, is_diff))
            if col not in ("Status", "ORIG_Name", "NEW_Name", "NAME_diff", "DIFF_Fields"):
                c.number_format = COORD_NUMBER_FORMAT
        ws.row_dimensions[ri].height = 18

    for ci, col in enumerate(cols, 1):
        sample = [str(col)] + [str(df.iloc[r, ci-1]) for r in range(min(300, len(df)))]
        w = min(max(len(v) + 2 for v in sample), 46)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ref = f"A1:{get_column_letter(len(cols))}{len(df)+1}"
    t = Table(displayName=table_name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)
    ws.freeze_panes = "B2"


def write_overview(wb, title, df_all, meta):
    """Write overview/summary sheet."""
    ws = wb.create_sheet(title)
    for col, w in {"A": 3, "B": 32, "C": 14, "D": 10, "E": 42}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:E2")
    ws["B2"] = "Measurement Points Comparison Report"
    ws["B2"].font = Font(name=FONT_NAME, size=16, bold=True, color=PALETTE["TITLE"])
    ws.row_dimensions[2].height = 28

    ws.merge_cells("B3:E3")
    ws["B3"] = f"Workbook: {meta['workbook']}   |   {datetime.now().strftime('%Y-%m-%d  %H:%M')}"
    ws["B3"].font = Font(name=FONT_NAME, size=10, italic=True, color="888888")
    ws.row_dimensions[3].height = 18

    ws.merge_cells("B4:E4")
    ws["B4"] = (f"Sheet 1: {meta['orig_sheet']} ({meta['n_orig']} rows)   |   "
                f"Sheet 2: {meta['new_sheet']} ({meta['n_new']} rows)   |   "
                f"XYZ tolerance: {meta['tol']} mm   |   "
                f"IJK: {'on' if meta['use_ijk'] else 'off'}")
    ws["B4"].font = Font(name=FONT_NAME, size=10, italic=True, color="AAAAAA")
    ws.row_dimensions[4].height = 16

    hrow = 6
    for ci, h in enumerate(["Status", "Count", "Color", "Description"], 2):
        ap(ws.cell(hrow, ci, h), hdr_style())
    ws.row_dimensions[hrow].height = 24

    vc = df_all["Status"].value_counts()
    for ri, code in enumerate(["MATCH","NAME_CHANGED","COORD_CHANGED","REPLACED?","DELETED","ADDED"], hrow+1):
        cnt = vc.get(code, 0)
        bg = PALETTE.get(code, "FFFFFF")
        ws.row_dimensions[ri].height = 22
        for ci, val in enumerate([code, cnt, "", STATUS_LABEL[code]], 2):
            c = ws.cell(ri, ci, val)
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name=FONT_NAME, size=11, bold=(ci in (2, 3)))
            c.border = thin()
            c.alignment = Alignment(horizontal="center" if ci in (2, 3, 4) else "left", vertical="center")

    tr = hrow + 6
    ws.row_dimensions[tr].height = 22
    for ci, val in enumerate(["TOTAL", len(df_all), "", "Total records"], 2):
        c = ws.cell(tr, ci, val)
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.border = thin()
        c.alignment = Alignment(horizontal="center" if ci in (2, 3, 4) else "left", vertical="center")

    lr = tr + 3
    ws.merge_cells(f"B{lr}:E{lr}")
    ws[f"B{lr}"] = "Orange text in diff columns means the value is not zero"
    ws[f"B{lr}"].font = Font(name=FONT_NAME, size=10, italic=True, color=PALETTE["DIFF_CELL"])
    ws.row_dimensions[lr].height = 18


def delete_report_sheets(wb, prefix):
    """Delete existing CMP_* sheets."""
    for suffix in REPORT_SUFFIXES:
        name = f"{prefix}{suffix}"
        if name in wb.sheetnames:
            del wb[name]


def write_close_points_sheet(wb, title, df_close, meta):
    """Write close points report sheet."""
    ws = wb.create_sheet(title)
    ws.merge_cells("A1:M1")
    ws["A1"] = f"Pairs of points within {meta['close_points_tol']:.3f} mm in 3D distance"
    ws["A1"].font = Font(name=FONT_NAME, size=14, bold=True, color=PALETTE["TITLE"])
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:M2")
    ws["A2"] = f"Original sheet: {meta['orig_sheet']}   |   New sheet: {meta['new_sheet']}"
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color="888888")
    ws.row_dimensions[2].height = 18

    if df_close.empty:
        ws.merge_cells("A4:M4")
        ws["A4"] = "No close point pairs were found within the XYZ tolerance radius."
        ws["A4"].font = Font(name=FONT_NAME, size=11, italic=True)
        return

    close_cols = list(df_close.columns)
    header_row = 4
    for ci, col in enumerate(close_cols, 1):
        ap(ws.cell(header_row, ci, col), hdr_style())
    ws.row_dimensions[header_row].height = 24

    for ri, (_, row) in enumerate(df_close.iterrows(), header_row + 1):
        status = str(row.get("Status", "REPLACED?"))
        for ci, col in enumerate(close_cols, 1):
            value = row[col]
            if isinstance(value, float) and np.isnan(value):
                value = ""
            cell = ws.cell(ri, ci, value)

            is_diff = col in ("NAME_diff", "dX", "dY", "dZ") and value not in ("", 0, 0.0)
            if isinstance(value, float) and is_diff and abs(value) < 1e-9:
                is_diff = False

            ap(cell, data_style(status, is_diff))
            if col.endswith(("_X", "_Y", "_Z")) or col in {"dX", "dY", "dZ", "Distance_3D"}:
                cell.number_format = COORD_NUMBER_FORMAT
        ws.row_dimensions[ri].height = 18

    for ci, col in enumerate(close_cols, 1):
        sample = [str(col)] + [str(df_close.iloc[r, ci - 1]) for r in range(min(300, len(df_close)))]
        width = min(max(len(v) + 2 for v in sample), 24)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ref = f"A{header_row}:{get_column_letter(len(close_cols))}{len(df_close) + header_row}"
    table = Table(displayName="ClosePoints", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = f"B{header_row + 1}"


def write_nearest_points_sheet(wb, title, df_nearest, meta):
    """Write nearest points report sheet."""
    ws = wb.create_sheet(title)
    ws.merge_cells("A1:M1")
    ws["A1"] = "Nearest new point for each original point"
    ws["A1"].font = Font(name=FONT_NAME, size=14, bold=True, color=PALETTE["TITLE"])
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:M2")
    ws["A2"] = f"Original sheet: {meta['orig_sheet']}   |   New sheet: {meta['new_sheet']}"
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color="888888")
    ws.row_dimensions[2].height = 18

    if df_nearest.empty:
        ws.merge_cells("A4:M4")
        ws["A4"] = "No nearest-point pairs could be computed (missing coordinates)."
        ws["A4"].font = Font(name=FONT_NAME, size=11, italic=True)
        return

    nearest_cols = list(df_nearest.columns)
    header_row = 4
    for ci, col in enumerate(nearest_cols, 1):
        ap(ws.cell(header_row, ci, col), hdr_style())
    ws.row_dimensions[header_row].height = 24

    for ri, (_, row) in enumerate(df_nearest.iterrows(), header_row + 1):
        for ci, col in enumerate(nearest_cols, 1):
            value = row[col]
            if isinstance(value, float) and np.isnan(value):
                value = ""
            cell = ws.cell(ri, ci, value)
            ap(cell, data_style("ADDED"))
            if col.endswith(("_X", "_Y", "_Z")) or col in {"dX", "dY", "dZ", "Distance_3D"}:
                cell.number_format = COORD_NUMBER_FORMAT
        ws.row_dimensions[ri].height = 18

    for ci, col in enumerate(nearest_cols, 1):
        sample = [str(col)] + [str(df_nearest.iloc[r, ci - 1]) for r in range(min(300, len(df_nearest)))]
        width = min(max(len(v) + 2 for v in sample), 24)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ref = f"A{header_row}:{get_column_letter(len(nearest_cols))}{len(df_nearest) + header_row}"
    table = Table(displayName="NearestPoints", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)
    ws.freeze_panes = f"B{header_row + 1}"


def write_report_sheets(wb, df_all, meta, prefix=REPORT_PREFIX):
    """Write all report sheets."""
    delete_report_sheets(wb, prefix)
    df_close = meta["df_close"]
    write_overview(wb, f"{prefix}Overview", df_all, meta)
    write_data_sheet(wb, f"{prefix}All Results", df_all, "AllResults")
    write_data_sheet(wb, f"{prefix}Match", df_all[df_all.Status=="MATCH"].reset_index(drop=True), "Match")
    write_data_sheet(wb, f"{prefix}Name Changed", df_all[df_all.Status=="NAME_CHANGED"].reset_index(drop=True), "NameChanged")
    write_data_sheet(wb, f"{prefix}Coord Changed", df_all[df_all.Status=="COORD_CHANGED"].reset_index(drop=True), "CoordChanged")
    write_data_sheet(wb, f"{prefix}Replaced", df_all[df_all.Status=="REPLACED?"].reset_index(drop=True), "Replaced")
    write_data_sheet(wb, f"{prefix}Deleted", df_all[df_all.Status=="DELETED"].reset_index(drop=True), "Deleted")
    write_data_sheet(wb, f"{prefix}Added", df_all[df_all.Status=="ADDED"].reset_index(drop=True), "Added")
    write_close_points_sheet(wb, f"{prefix}Close Points", df_close, meta)


def sanitize_external_query_metadata(wb):
    """Convert query tables to regular tables."""
    for ws in wb.worksheets:
        for table in ws._tables.values():
            if getattr(table, "tableType", None) == "queryTable":
                table.tableType = "worksheet"
                for column in table.tableColumns:
                    if hasattr(column, "queryTableFieldId"):
                        column.queryTableFieldId = None


EXCEL_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def clean_workbook_xml_bytes(data):
    """Clean workbook.xml of ExternalData references."""
    root = ET.fromstring(data)
    defined_names = root.find(f"{{{EXCEL_MAIN_NS}}}definedNames")
    if defined_names is None:
        return data
    for defined_name in list(defined_names):
        if defined_name.attrib.get("name", "").startswith("ExternalData_"):
            defined_names.remove(defined_name)
    if not list(defined_names):
        root.remove(defined_names)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def clean_table_xml_bytes(data):
    """Clean table.xml of query table attributes."""
    root = ET.fromstring(data)
    if root.attrib.get("tableType") == "queryTable":
        root.attrib["tableType"] = "worksheet"
    for column in root.findall(f".//{{{EXCEL_MAIN_NS}}}tableColumn"):
        column.attrib.pop("queryTableFieldId", None)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def clean_saved_workbook_package(workbook_path):
    """Clean workbook package of query metadata."""
    workbook_path = Path(workbook_path)
    with NamedTemporaryFile(delete=False, suffix=workbook_path.suffix) as tmp:
        temp_path = Path(tmp.name)
    with ZipFile(workbook_path, "r") as src_zip, ZipFile(temp_path, "w") as dst_zip:
        for info in src_zip.infolist():
            data = src_zip.read(info.filename)
            if info.filename == "xl/workbook.xml":
                data = clean_workbook_xml_bytes(data)
            elif info.filename.startswith("xl/tables/") and info.filename.endswith(".xml"):
                data = clean_table_xml_bytes(data)
            dst_zip.writestr(info, data)
    shutil.move(str(temp_path), str(workbook_path))


def save_workbook_clean(wb, workbook_path):
    """Save workbook with cleaning."""
    sanitize_external_query_metadata(wb)
    wb.save(workbook_path)
    clean_saved_workbook_package(workbook_path)


def find_existing_append_block(ws, headers, legacy_headers=None, header_row=None):
    """Find existing append block in original sheet."""
    row_values = []
    if header_row is None:
        header_row = 1
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_idx).value
        row_values.append("" if value is None else str(value).strip())
    header_sets = [headers]
    if legacy_headers:
        header_sets.append(legacy_headers)
    for header_set in header_sets:
        block_len = len(header_set)
        for start_idx in range(0, len(row_values) - block_len + 1):
            if row_values[start_idx:start_idx + block_len] == header_set:
                return start_idx + 1
    return None


def disable_sheet_filter(ws):
    """Remove sheet filter and unhide rows."""
    if getattr(ws.auto_filter, "ref", None):
        ws.auto_filter.ref = None
    for row_idx in range(1, ws.max_row + 1):
        if row_idx in ws.row_dimensions:
            ws.row_dimensions[row_idx].hidden = False


def format_original_coordinate_columns(ws, header_row, data_start_row, source_end_col):
    """Format coordinate columns with number format."""
    source_coord_columns = {}
    for col_idx in range(1, source_end_col + 1):
        header_value = ws.cell(header_row, col_idx).value
        if header_value is None:
            continue
        normalized = str(header_value).strip()
        if normalized in COORD_FIELDS and normalized not in source_coord_columns:
            source_coord_columns[normalized] = col_idx
    for col_idx in source_coord_columns.values():
        for row_idx in range(data_start_row, ws.max_row + 1):
            ws.cell(row_idx, col_idx).number_format = COORD_NUMBER_FORMAT


def clear_existing_output_rows(ws, data_start_row, original_data_rows, source_end_col):
    """Clear existing output rows after original data."""
    row_idx = data_start_row + original_data_rows
    while row_idx <= ws.max_row:
        has_original_data = any(
            ws.cell(row_idx, col_idx).value not in (None, "")
            for col_idx in range(1, source_end_col + 1)
        )
        if has_original_data:
            break
        ws.delete_rows(row_idx, 1)


def apply_group_labels(ws, style_kit, source_end_col, status_col, new_start_col, new_end_col, diff_start_col, diff_end_col):
    """Apply group labels to row 1."""
    for merge_range in list(ws.merged_cells.ranges):
        if merge_range.min_row == 1 and merge_range.max_row == 1:
            ws.unmerge_cells(str(merge_range))

    def write_group(start_col, end_col, label):
        if start_col > end_col:
            return
        if start_col == end_col:
            cell = ws.cell(1, start_col, label)
            ap(cell, original_sheet_group_style(style_kit, label))
            return
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(1, start_col, label)
        ap(cell, original_sheet_group_style(style_kit, label))

    from .schema import STATUS_GROUP_LABEL
    write_group(1, source_end_col, "Original data")
    cell = ws.cell(1, status_col, STATUS_GROUP_LABEL)
    ap(cell, original_sheet_group_style(style_kit, STATUS_GROUP_LABEL))
    write_group(new_start_col, new_end_col, "New Data")
    write_group(diff_start_col, diff_end_col, "Difference")


def append_results_to_original_sheet(wb, sheet_name, df_all, n_orig):
    """Append comparison results to original sheet."""
    ws = wb[sheet_name]
    disable_sheet_filter(ws)

    header_row = 1
    if any(ws.cell(1, c).value and str(ws.cell(1, c).value).strip() in GROUP_HEADER_LABELS for c in range(1, ws.max_column + 1)):
        header_row = 2
    else:
        ws.insert_rows(1)
        header_row = 2

    headers = [header for _, header in ORIGINAL_SHEET_APPEND_MAP]
    start_col = find_existing_append_block(ws, headers, legacy_headers=LEGACY_ORIGINAL_SHEET_APPEND_HEADERS, header_row=header_row)
    source_end_col = start_col - 1 if start_col is not None else ws.max_column
    if start_col is None:
        start_col = source_end_col + 1

    end_col = start_col + len(headers) - 1
    data_start_row = header_row + 1
    style_kit = infer_original_sheet_style_kit(ws, header_row, data_start_row, source_end_col)
    format_original_coordinate_columns(ws, header_row, data_start_row, source_end_col)

    for col_idx in range(start_col, end_col + 1):
        for row_idx in range(header_row, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.alignment = Alignment(horizontal="general", vertical="bottom")
            cell.font = Font(name=FONT_NAME, size=10)

    clear_existing_output_rows(ws, data_start_row, n_orig, source_end_col)

    for offset, (_, header) in enumerate(ORIGINAL_SHEET_APPEND_MAP):
        cell = ws.cell(header_row, start_col + offset, header)
        fill = PatternFill(patternType="solid", fgColor="FFEDEDED") if header == "STATUS" else None
        ap(cell, original_sheet_header_style(style_kit, fill=fill))

    orig_rows = df_all.iloc[:n_orig].reset_index(drop=True)
    for row_idx, (_, result_row) in enumerate(orig_rows.iterrows(), start=data_start_row):
        for offset, (source_col, _) in enumerate(ORIGINAL_SHEET_APPEND_MAP):
            value = result_row[source_col] if source_col in result_row else ""
            if isinstance(value, float) and np.isnan(value):
                value = ""
            is_diff = source_col.endswith("_diff") or source_col == "DIFF_Fields"
            cell = ws.cell(row_idx, start_col + offset, value)
            if source_col == "Status":
                ap(cell, original_sheet_data_style(style_kit, is_diff=False,
                    fill=STATUS_FILLS.get(str(value), PatternFill(patternType="solid", fgColor=GROUP_FILL_ORIGINAL.fgColor)),
                    bold=True, horizontal="center"))
            else:
                ap(cell, original_sheet_data_style(style_kit, is_diff=is_diff and value not in ("", 0, 0.0)))
            if source_col in ORIGINAL_SHEET_NUMERIC_COLUMNS:
                cell.number_format = COORD_NUMBER_FORMAT

    added_rows = df_all[df_all.Status == "ADDED"].reset_index(drop=True)
    append_row_idx = data_start_row + n_orig
    for _, added_row in added_rows.iterrows():
        for offset, (source_col, _) in enumerate(ORIGINAL_SHEET_APPEND_MAP):
            value = ""
            if source_col == "Status":
                value = "NEW"
            elif source_col.startswith("NEW_"):
                value = added_row.get(source_col, "")
                if isinstance(value, float) and np.isnan(value):
                    value = ""
            cell = ws.cell(append_row_idx, start_col + offset, value)
            is_diff = source_col.endswith("_diff") or source_col == "DIFF_Fields"
            if source_col == "Status":
                ap(cell, original_sheet_data_style(style_kit, is_diff=False, fill=STATUS_FILLS["NEW"],
                    bold=True, horizontal="center"))
            else:
                ap(cell, original_sheet_data_style(style_kit, is_diff=is_diff and value not in ("", 0, 0.0)))
            if source_col in ORIGINAL_SHEET_NUMERIC_COLUMNS:
                cell.number_format = COORD_NUMBER_FORMAT
        append_row_idx += 1

    for offset, (_, header) in enumerate(ORIGINAL_SHEET_APPEND_MAP):
        col_idx = start_col + offset
        sample_values = [str(header)]
        sample_values.extend(str(ws.cell(row_idx, col_idx).value)
            for row_idx in range(data_start_row, ws.max_row + 1)
            if ws.cell(row_idx, col_idx).value not in (None, ""))
        width = min(max((len(value) + 2) for value in sample_values), 24)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    apply_group_labels(ws, style_kit, source_end_col=source_end_col, status_col=start_col,
        new_start_col=start_col + 1, new_end_col=start_col + 7, diff_start_col=start_col + 8, diff_end_col=end_col)


def run_comparison(workbook_path, orig_sheet, new_sheet, tol, ijk_tol, use_ijk,
                   create_report_sheets=True, close_points_tol=None):
    """Run comparison and write results to a new workbook file."""
    if close_points_tol is None:
        close_points_tol = DEFAULTS["close_points_tol"]

    src_path = Path(workbook_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = src_path.parent / f"{src_path.stem}_CMP_{timestamp}{src_path.suffix}"

    df_orig_raw = read_sheet(workbook_path, orig_sheet)
    df_new_raw = read_sheet(workbook_path, new_sheet)
    df_orig = prepare_sheet(df_orig_raw, orig_sheet)
    df_new = prepare_sheet(df_new_raw, new_sheet)

    df_all = compare(df_orig, df_new, tol, ijk_tol, use_ijk)
    df_all, moved_pairs = link_moved_pairs(df_all, df_orig, df_new, close_points_tol)
    df_close = build_close_points_report(df_orig, df_new, close_points_tol, df_all)

    meta = {
        "workbook": Path(workbook_path).name,
        "orig_sheet": orig_sheet,
        "new_sheet": new_sheet,
        "n_orig": len(df_orig),
        "n_new": len(df_new),
        "tol": tol,
        "ijk_tol": ijk_tol,
        "use_ijk": use_ijk,
        "df_orig": df_orig,
        "df_new": df_new,
        "df_close": df_close,
        "moved_pairs": moved_pairs,
        "close_points_tol": close_points_tol,
    }

    shutil.copy2(workbook_path, output_path)

    keep_vba = Path(workbook_path).suffix.lower() == ".xlsm"
    wb = load_workbook(str(output_path), keep_vba=keep_vba)
    append_results_to_original_sheet(wb, orig_sheet, df_all, len(df_orig))
    if create_report_sheets:
        write_report_sheets(wb, df_all, meta)
    save_workbook_clean(wb, str(output_path))

    return {
        "df_all": df_all,
        "meta": meta,
        "orig_columns": list(df_orig_raw.columns),
        "new_columns": list(df_new_raw.columns),
        "create_report_sheets": create_report_sheets,
        "close_points_count": len(df_close),
        "moved_count": len(moved_pairs),
        "output_path": str(output_path),
    }
