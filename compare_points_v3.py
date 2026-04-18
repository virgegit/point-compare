#!/usr/bin/env python3
"""
compare_points_v3.py
====================
Compares two worksheets in one Excel workbook with measurement points
(Name, X, Y, Z, I, J, K).
Output: styled Excel sheets written back into the same workbook.

Change-categories
-----------------
  MATCH          – same name AND coordinates within tolerance
  NAME_CHANGED   – same coordinates, different name  → shows old/new name
  COORD_CHANGED  – same name, different coords       → shows dX/dY/dZ + dI/dJ/dK
  DELETED        – only in File 1  (by both name and coords)
  ADDED          – only in File 2  (by both name and coords)
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from point_compare.schema import LOGICAL_FIELDS
from point_compare.excel_io import read_sheet, prepare_sheet, run_comparison

# ════════════════════════════════════════════════════════════════
#  CONFIGURATION  – edit this section before running
# ════════════════════════════════════════════════════════════════
CONFIG = {
    # ── Input workbook ──────────────────────────────────────────
    "workbook_file": "PointsCompare_Template.xlsm",
    "sheet1": "☕ Sheet 1",
    "sheet2": "☕ Sheet 2",
    "report_prefix": "CMP_",

    # ── Tolerances ────────────────────────────────────────────
    "coord_tol": 0.05,
    "ijk_tol":   0.001,
    "compare_ijk": False,

    # ── Colour Palette (RRGGBB hex, no #) ─────────────────────
    "palette": {
        "MATCH":         "C6EFCE",
        "NAME_CHANGED":  "FFEB9C",
        "COORD_CHANGED": "FFD7D7",
        "DELETED":       "F4CCCC",
        "ADDED":         "D9EAD3",
        "DIFF_CELL":     "FF6600",
        "HEADER_BG":     "4472C4",
        "HEADER_FG":     "FFFFFF",
        "OVERVIEW_TITLE":"1F3864",
    },
}
# ════════════════════════════════════════════════════════════════


STATUS_LABEL = {
    "MATCH":         "✔ Match",
    "NAME_CHANGED":  "✎ Name Changed",
    "COORD_CHANGED": "⚠ Coordinates Changed",
    "DELETED":       "✖ Deleted (only in sheet 1)",
    "ADDED":         "＋ Added (only in sheet 2)",
}

REPORT_SUFFIXES = [
    "Overview",
    "All Results",
    "Match",
    "Name Changed",
    "Coord Changed",
    "Deleted",
    "Added",
]


def map_result_columns(df_all):
    """Map interactive.py's ORIG_*/NEW_* columns to v3's F1_*/F2_* convention."""
    mapping = {}
    for logical_field in LOGICAL_FIELDS:
        mapping[f"ORIG_{logical_field}"] = f"F1_{logical_field}"
        mapping[f"NEW_{logical_field}"] = f"F2_{logical_field}"
    return df_all.rename(columns=mapping)


FONT_NAME = "Calibri"


def thin_border(color="CCCCCC"):
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


class StyleKit:
    def __init__(self, palette):
        self.p = palette

    def hdr(self):
        return dict(
            font=Font(name=FONT_NAME, bold=True, color=self.p["HEADER_FG"], size=11),
            fill=PatternFill("solid", fgColor=self.p["HEADER_BG"]),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=thin_border(),
        )

    def data(self, status):
        return dict(
            font=Font(name=FONT_NAME, size=10),
            fill=PatternFill("solid", fgColor=self.p.get(status, "FFFFFF")),
            alignment=Alignment(horizontal="left", vertical="center"),
            border=thin_border(),
        )

    def diff(self, status):
        d = self.data(status)
        d["font"] = Font(name=FONT_NAME, size=10, bold=True, color=self.p["DIFF_CELL"])
        return d


def apply_style(cell, style_dict):
    for k, v in style_dict.items():
        setattr(cell, k, v)


DIFF_COLS = {"DIFF_Fields", "NAME_diff",
             "X_diff", "Y_diff", "Z_diff",
             "I_diff", "J_diff", "K_diff"}

ORDERED_COLS = [
    "Status",
    "F1_Name", "F1_X", "F1_Y", "F1_Z", "F1_I", "F1_J", "F1_K",
    "F2_Name", "F2_X", "F2_Y", "F2_Z", "F2_I", "F2_J", "F2_K",
    "NAME_diff", "X_diff", "Y_diff", "Z_diff", "I_diff", "J_diff", "K_diff",
    "DIFF_Fields",
]


def write_data_sheet(wb, title, df, table_name, sk):
    ws = wb.create_sheet(title)
    if df.empty:
        ws.cell(1, 1, "No data").font = Font(italic=True)
        return

    cols = [c for c in ORDERED_COLS if c in df.columns]
    for ci, col in enumerate(cols, 1):
        apply_style(ws.cell(1, ci, col), sk.hdr())
    ws.row_dimensions[1].height = 30

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        status = str(row.get("Status", ""))
        for ci, col in enumerate(cols, 1):
            val = row[col]
            if isinstance(val, float) and np.isnan(val):
                val = ""
            c = ws.cell(ri, ci, val)
            is_diff = col in DIFF_COLS
            is_meaningful = (is_diff and val not in ("", "–", 0, 0.0)
                             and not (isinstance(val, float) and abs(val) < 1e-9))
            apply_style(c, sk.diff(status) if is_meaningful else sk.data(status))
        ws.row_dimensions[ri].height = 18

    for ci, col in enumerate(cols, 1):
        sample = [str(col)] + [str(df.iloc[r, ci-1]) for r in range(min(300, len(df)))]
        w = min(max(len(v) + 2 for v in sample), 48)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ref = f"A1:{get_column_letter(len(cols))}{len(df)+1}"
    t = Table(displayName=table_name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(t)
    ws.freeze_panes = "B2"


def write_overview(wb, title, df_all, cfg, sk):
    ws = wb.create_sheet(title)
    p = sk.p

    for col, w in {"A": 3, "B": 30, "C": 12, "D": 10, "E": 38, "F": 22}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:F2")
    ws["B2"] = "Measurement Points Comparison Report"
    ws["B2"].font = Font(name=FONT_NAME, size=16, bold=True, color=p["OVERVIEW_TITLE"])
    ws.row_dimensions[2].height = 26

    ws.merge_cells("B3:F3")
    ws["B3"] = (f"Workbook: {cfg['workbook_file']}  |  "
                f"{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws["B3"].font = Font(name=FONT_NAME, size=10, italic=True, color="888888")

    ws.merge_cells("B4:F4")
    ws["B4"] = (f"Sheet 1: {cfg['sheet1']}  |  Sheet 2: {cfg['sheet2']}  |  "
                f"XYZ tolerance: {cfg['coord_tol']} mm  |  "
                f"IJK tolerance: {cfg['ijk_tol']}  |  "
                f"Compare IJK: {'Yes' if cfg['compare_ijk'] else 'No'}")
    ws["B4"].font = Font(name=FONT_NAME, size=10, italic=True, color="AAAAAA")

    hrow = 6
    for ci, h in enumerate(["Status", "Count", "Color", "Description"], 2):
        apply_style(ws.cell(hrow, ci, h), sk.hdr())
    ws.row_dimensions[hrow].height = 24

    vc = df_all["Status"].value_counts()
    for ri, code in enumerate(["MATCH", "NAME_CHANGED", "COORD_CHANGED", "DELETED", "ADDED"], hrow + 1):
        cnt = vc.get(code, 0)
        ws.row_dimensions[ri].height = 20
        for ci, val in enumerate([code, cnt, "", STATUS_LABEL[code]], 2):
            c = ws.cell(ri, ci, val)
            c.fill = PatternFill("solid", fgColor=p.get(code, "FFFFFF"))
            c.border = thin_border()
            c.font = Font(name=FONT_NAME, size=11, bold=(ci in (2, 3)))
            c.alignment = Alignment(horizontal="center" if ci in (2, 3, 4) else "left",
                                    vertical="center")

    tr = hrow + 6
    ws.row_dimensions[tr].height = 20
    for ci, val in enumerate(["TOTAL", len(df_all), "", "Total records"], 2):
        c = ws.cell(tr, ci, val)
        c.border = thin_border()
        c.font = Font(name=FONT_NAME, size=11, bold=True)
        c.alignment = Alignment(horizontal="center" if ci in (2, 3, 4) else "left",
                                 vertical="center")

    # Palette reference
    pr = tr + 3
    ws.merge_cells(f"B{pr}:F{pr}")
    ws[f"B{pr}"] = "Color Palette"
    ws[f"B{pr}"].font = Font(name=FONT_NAME, size=12, bold=True)
    ws.row_dimensions[pr].height = 20
    pr += 1
    for ci, h in enumerate(["Key", "Hex", "Sample"], 2):
        apply_style(ws.cell(pr, ci, h), sk.hdr())
    for ri2, (k, v) in enumerate(p.items(), pr + 1):
        ws.row_dimensions[ri2].height = 18
        ws.cell(ri2, 2, k).border = thin_border()
        ws.cell(ri2, 3, v).border = thin_border()
        sc = ws.cell(ri2, 4, "")
        sc.fill = PatternFill("solid", fgColor=v)
        sc.border = thin_border()

    mr = pr + len(p) + 3
    ws.merge_cells(f"B{mr}:F{mr}")
    ws[f"B{mr}"] = "Required source columns: Name, X, Y, Z, I, J, K"
    ws[f"B{mr}"].font = Font(name=FONT_NAME, size=11, italic=True)
    ws.row_dimensions[mr].height = 20


def build_report(cfg):
    print("Comparing points …")

    workbook_file = cfg["workbook_file"]
    result = run_comparison(
        workbook_file,
        cfg["sheet1"],
        cfg["sheet2"],
        cfg["coord_tol"],
        cfg["ijk_tol"],
        cfg["compare_ijk"],
        create_report_sheets=False,
    )

    df_all = result["df_all"]

    # Map column names from interactive convention to v3 convention
    df_all = map_result_columns(df_all)

    sk = StyleKit(cfg["palette"])
    vc = df_all["Status"].value_counts()

    out = Path(workbook_file)
    wb = load_workbook(out, keep_vba=(out.suffix.lower() == ".xlsm"))
    for suffix in REPORT_SUFFIXES:
        name = f"{cfg['report_prefix']}{suffix}"
        if name in wb.sheetnames:
            del wb[name]

    write_overview(wb, f"{cfg['report_prefix']}Overview", df_all, cfg, sk)
    write_data_sheet(wb, f"{cfg['report_prefix']}All Results",  df_all,                          "AllResults",   sk)
    write_data_sheet(wb, f"{cfg['report_prefix']}Match",        df_all[df_all.Status=="MATCH"].reset_index(drop=True),  "Match",        sk)
    write_data_sheet(wb, f"{cfg['report_prefix']}Name Changed", df_all[df_all.Status=="NAME_CHANGED"].reset_index(drop=True),  "NameChanged",  sk)
    write_data_sheet(wb, f"{cfg['report_prefix']}Coord Changed", df_all[df_all.Status=="COORD_CHANGED"].reset_index(drop=True), "CoordChanged", sk)
    write_data_sheet(wb, f"{cfg['report_prefix']}Deleted",      df_all[df_all.Status=="DELETED"].reset_index(drop=True),        "Deleted",      sk)
    write_data_sheet(wb, f"{cfg['report_prefix']}Added",        df_all[df_all.Status=="ADDED"].reset_index(drop=True),          "Added",        sk)

    wb.save(out)

    print(f"\n Report saved in source workbook → {out}")
    print("\n=== Summary ===")
    for s in ["MATCH", "NAME_CHANGED", "COORD_CHANGED", "DELETED", "ADDED"]:
        print(f"  {STATUS_LABEL[s]}: {vc.get(s, 0)}")
    print(f"  Total: {len(df_all)}")
    return df_all


if __name__ == "__main__":
    build_report(CONFIG)
